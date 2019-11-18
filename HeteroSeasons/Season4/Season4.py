#Andrew Anastasiades
#This Code narrows down the possible perfect matches from the MTV show "Are you the one"

from itertools import permutations
import openpyxl

# total_matches = list(permutations([1,2,3,4,5,6,7,8,9,10]))
#The position in tuple ceremony_x represents a guy (indexed 0-9)
#The number in that position represesnts a girl (indexed 1-10)
total_matches = list()
week_1_possible = list()
week_2_possible = list()
week_3_possible = list()
week_4_possible = list()
week_5_possible = list()
week_6_possible = list()
week_7_possible = list()
week_8_possible = list()
week_9_possible = list()
week_10_possible = list()

def get_total(pairs):
  totes = list(permutations(range(1,pairs+1)))
  print("Total matches is ", len(totes))
  return totes


total_matches_filename = "All Possible Permutations.txt"

#Week 1
week_1_booth = (0,0,0,0,0,0,0,0,0,8,0)
week_1_booth_match = False
week_1_ceremony = (7,3,5,10,1,11,4,8,6,9,2)
week_1_matches = 3
week_1_filename = "Week 1 Possibilities.txt"

#Week 2
week_2_booth = (1,0,0,0,0,0,0,0,0,0,0)
week_2_booth_match = False
week_2_ceremony = (9, 1, 5, 10, 7, 4, 2, 8, 6, 3, 11)
week_2_matches = 3
week_2_filename = "Week 2 Possibilities.txt"

#Week 3
week_3_booth = (0,0,0,0,0,0,0,8,0,0,0)
week_3_booth_match = True
week_3_ceremony = (7,3,5,11,4,1,10,8,6,2,9)
week_3_matches = 3
week_3_filename = "Week 3 Possibilities.txt"

#Week 4
week_4_booth = (0,6,0,0,0,0,0,0,0,0,0)
week_4_booth_match = False
week_4_ceremony = (10,5,1,7,2,6,11,8,4,3,9)
week_4_matches = 2
week_4_filename = "Week 4 Possibilities.txt"

#Week 5
week_5_booth = (0,0,0,0,0,0,0,0,0,0,2)
week_5_booth_match = False
week_5_ceremony = (11,10,5,2,4,7,1,8,6,3,9)
week_5_matches = 4
week_5_filename = "Week 5 Possibilities.txt"

#Week 6
week_6_booth = (0,3,0,0,0,0,0,0,0,0,0)
week_6_booth_match = False
week_6_ceremony = (7,10,5,11,1,2,4,8,6,3,9)
week_6_matches = 4
week_6_filename = "Week 6 Possibilities.txt"

#Week 7
week_7_booth = (0,0,0,0,0,0,0,0,0,0,10)
week_7_booth_match = False
week_7_ceremony = (7,10,3,11,1,2,4,8,6,5,9)
week_7_matches = 4
week_7_filename = "Week 7 Possibilities.txt"

#Week 8
week_8_booth = (0,0,0,0,0,0,0,0,6,0,0)
week_8_booth_match = True
week_8_ceremony = (11,10,7,1,4,2,5,8,6,9,3)
week_8_matches = 4
week_8_filename = "Week 8 Possibilities.txt"

#Week 9
week_9_booth = (0,0,11,0,0,0,0,0,0,0,0)
week_9_booth_match = False
week_9_ceremony = (3,2,9,7,10,1,5,8,6,4,11)
week_9_matches = 4
week_9_filename ="Week 9 Possibilities.txt"

#Week 10
week_10_booth = (0,10,0,0,0,0,0,0,0,0,0)
week_10_booth_match = True
week_10_ceremony = (3,10,2,4,7,11,5,8,6,1,9)
week_10_matches = 11
week_10_filename = "Week 10 Possibilities.txt"

# Weekly info stored as a dict
#   Key = "Week #"
#   Value is a list [week_string, booth, booth_match, ceremony, ceremony_matches, starting_set, ending_set, week_filename]

week_info = {
"Week 0" : ["Week 0", "No Booth", "No booth match", "No ceremony", "no ceremony matches", "No starting set", total_matches, total_matches_filename],
"Week 1" : ["Week 1", week_1_booth, week_1_booth_match, week_1_ceremony, week_1_matches, total_matches, week_1_possible, week_1_filename],
"Week 2" : ["Week 2", week_2_booth, week_2_booth_match, week_2_ceremony, week_2_matches, week_1_possible, week_2_possible, week_2_filename],
"Week 3" : ["Week 3", week_3_booth, week_3_booth_match, week_3_ceremony, week_3_matches, week_2_possible, week_3_possible, week_3_filename],
"Week 4" : ["Week 4", week_4_booth, week_4_booth_match, week_4_ceremony, week_4_matches, week_3_possible, week_4_possible, week_4_filename],
"Week 5" : ["Week 5", week_5_booth, week_5_booth_match, week_5_ceremony, week_5_matches, week_4_possible, week_5_possible, week_5_filename],
"Week 6" : ["Week 6", week_6_booth, week_6_booth_match, week_6_ceremony, week_6_matches, week_5_possible, week_6_possible, week_6_filename],
"Week 7" : ["Week 7", week_7_booth, week_7_booth_match, week_7_ceremony, week_7_matches, week_6_possible, week_7_possible, week_7_filename],
"Week 8" : ["Week 8", week_8_booth, week_8_booth_match, week_8_ceremony, week_8_matches, week_7_possible, week_8_possible, week_8_filename],
"Week 9" : ["Week 9", week_9_booth, week_9_booth_match, week_9_ceremony, week_9_matches, week_8_possible, week_9_possible, week_9_filename],
"Week 10" : ["Week 10", week_10_booth, week_10_booth_match, week_10_ceremony, week_10_matches, week_9_possible, week_10_possible, week_10_filename]
}


guys = {
  0 : ["Andrew", dict()],
  1 : ["Brett", dict()],
  2 : ["Cam", dict()],
  3 : ["Daniel", dict()],
  4 : ["Kwasi", dict()],
  5 : ["Lewis", dict()],
  6 : ["Moe", dict()],
  7 : ["Shamoy", dict()],
  8 : ["Tevin", dict()],
  9 : ["Tomas", dict()],
  10 : ["Zak", dict()]
}

girls = {
  1 : "Asia",
  2 : "Bria",
  3 : "Cali",
  4 : "Jasmine",
  5 : "Kayla",
  6 : "Kenya",
  7 : "Lauren",
  8 : "Maria",
  9 : "Morgan",
  10 : "Nutsa",
  11 : "Sam"
}

# The following functions perform file IO and the essential selection process:
#   make_tuple(week_key, dictionary)
#   narrow_down(week_key, dictionary)
#   write_to_file(week_key)
#   write_to_workbook(week_key):

def make_tuple(week_key, dictionary = week_info):
  filename = dictionary[week_key][7]
  ending_set = dictionary[week_key][6]
  ending_set = list()
  with open(filename, mode='r', encoding='utf=8') as f:
      for line in f:
          line = line[1:34] #they all have a '(' at the beginning and end
          line_string_list = line.split(', ')
          for i in range(len(line_string_list)):
              line_string_list[i] = int(line_string_list[i])
          line_int_tuple = tuple(line_string_list)
          ending_set.append(line_int_tuple)
  return ending_set

def narrow_down(week_key, dictionary = week_info):
  #pull the relevant info from week_info dictionary
  week_string = dictionary[week_key][0]
  booth = dictionary[week_key][1]
  booth_match = dictionary[week_key][2]
  ceremony = dictionary[week_key][3]
  ceremony_matches = dictionary[week_key][4]
  starting_set = dictionary[week_key][5]
  ending_set = dictionary[week_key][6]
  filename = dictionary[week_key][7]

  for match in starting_set:
    count=0
    perfect_match = 0

    for i in range(11):
      if match[i] == ceremony[i]:
        count+=1
      if match[i] == booth[i]:
        perfect_match+=1  
    if ((count == ceremony_matches) and (perfect_match == booth_match)):
      ending_set.append(match)
  print("Matches possible after", week_string, len(ending_set))

def write_to_file(week_key):
  filename = week_info[week_key][7]
  with open(filename, mode='w', encoding='utf-8') as f:
      for each in week_info[week_key][6]:
          f.write(str(each))
          f.write('\n')

def write_to_workbook(week_key):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = week_key + " Possible Matches"
  testing_set = week_info[week_key][6]
  for i in range(len(testing_set)):
    this_match = testing_set[i]
    for j in range(len(this_match)):
      this_cell = ws.cell(i+1, j+1, this_match[j])
  wb.save(week_key + ".xlsx")

# This function executes all the weekly essentional functions
def weekly_update(week_key, previous_week_key, dictionary = week_info):
  week_info[week_key][5] = make_tuple(previous_week_key)
  narrow_down(week_key)
  write_to_file(week_key)
  write_to_workbook(week_key)

#To conserve computation time, I write the each week's possible scenarios to file
#After week 1, this is a 4 step process:
# 1) Read file from previous week and clean it up as a list of integer tuples
# 2) Assign this list to the current week's entry in the dictionary 'week_info'
# 3) Perform analysis on the most recent information
# 4) Write this information to a file

#total_matches = get_total(11)

#weekly_update("Week 1", "Week 0")
#weekly_update("Week 2", "Week 1")
#weekly_update("Week 3", "Week 2")
#weekly_update("Week 4", "Week 3")
#weekly_update("Week 5", "Week 4")
#weekly_update("Week 6", "Week 5")
#weekly_update("Week 7", "Week 6")
#weekly_update("Week 8", "Week 7")
#weekly_update("Week 9", "Week 8")
weekly_update("Week 10", "Week 9")

def list_possible():
  for guy in range(11):
    guys_matches = guys.get(guy)[1]
    guys_name = guys.get(guy)[0]
    print(guys_name, "'s matches are") 
    print(guys_matches)

test_dict = {
  "test_start" : list(),
  "test_end" : list(),
  "test_ceremony" : (10,6,8,2,5,1,3,7,4,9,11),
  "test_booth" : (0,0,0,0,0,0,0,0,0,0,0),
  "test_match" : False
}

def test_set(matches, dictionary = test_dict):
  start_set = dictionary["test_start"]
  end_set = list()
  ceremony = dictionary["test_ceremony"]
  booth = dictionary["test_booth"]
  match_test = dictionary["test_match"]

  for match in start_set:
    count=0
    perfect_match = 0

    for i in range(10):
      if match[i] == ceremony[i]:
        count+=1
      if match[i] == booth[i]:
        perfect_match+=1  
    if ((count == matches) and (perfect_match == match_test)):
      end_set.append(match)
  print(matches, "matches gives ", len(end_set), "possibilities")

def show_possible(week_key, dictionary = week_info):
  possibilities = dictionary.get(week_key)[6]
  for each in possibilities:
    print(each)

def named_matches(week_key):
  for each in week_info[week_key][6]:
    for i in range(10):
      print(guys[i][0], girls[each[i]])
    print("")

def weekly_progress():
  for each in week_info.keys():
    filename = week_info[each][7]
    try:
      with open(filename, mode='r', encoding='utf-8') as f:
          print(week_info[each][0], "has", len(f.readlines()), "possibilities")
    except:
      print(week_info[each][0], "is still unknown")

# The functions in the next section are related to analyzing the matches remaining
# at the end of a particular week
def clear_guy_dict():
  for guy in guys.keys():
    for i in range(1,12):
      guys[guy][1][i]=0
  print("Guys options cleared")

def count_guys_matches(week_key):
  ending_set = week_info[week_key][6]
  for each in ending_set:
    for i in range(len(each)):
      guys[i][1][each[i]]+=1

def show_guys_matches():
  for guy in guys.keys():
    temp_list = []
    for girl_count in guys[guy][1].values():
      temp_list.append(girl_count)
    print(guys[guy][0], "is matched with", temp_list)

def show_guys_best_matches(week_key):
  week_total = len(week_info[week_key][6])
  for guy in guys.keys():
    temp_list = []
    for girl_count in guys[guy][1].values():
      temp_list.append(girl_count)
    sorted_temp_list = sorted(temp_list)
    best = sorted_temp_list[-1]
    second = sorted_temp_list[-2]
    best_girl = girls[temp_list.index(best)+1]
    second_girl = girls[temp_list.index(second)+1]
    print(guys[guy][0], "is best matched with", best_girl, "({0:.0%})".format(best/week_total), "then" ,second_girl, "({0:.0%})".format(second/week_total) )

def who_is_not_a_match():
  print("")
  for i in guys.keys():
    this_guys_match = list()
    match_names = str()
    for k in range(len(guys[i][1])):
      if guys[i][1][k+1] == 0:
        this_guys_match.append(k+1)
    for each in this_guys_match:
      match_names+=(" "+ girls[each])
    print(guys[i][0] + " is not a match with" + match_names)
    
  
def best_match_suite(week_key):
  clear_guy_dict()
  count_guys_matches(week_key)
  show_guys_matches()
  print("")
  print("The Best Matches of", week_key)
  show_guys_best_matches(week_key)
  who_is_not_a_match()

# The following funtions allow us to analyze the possible match-ups for a particular week:
#   match_distribution(week_key)
#   write_test_to_workbook(match_distribution, week_key, test_filename = "Match Distribution"):

def match_distribution(week_key):
  test_results = list()
  testing_set = week_info[week_key][6]
  for comparator in testing_set:
    comparator_list = [0,0,0,0,0,0,0,0,0,0,0,0]
    for each in testing_set:
        count = 0
        for i in range(11):
          if each[i] == comparator[i]:
            count+=1
        comparator_list[count]+=1
    test_results.append(comparator_list)
  return test_results

def write_test_to_workbook(match_distribution, week_key, test_filename = "Match Distribution"):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = week_key + " Match Distribution"
  testing_set = match_distribution
  for i in range(len(testing_set)):
    this_match = testing_set[i]
    for j in range(len(this_match)):
      this_cell = ws.cell(i+1, j+1, this_match[j])
  wb.save(test_filename + ".xlsx")

# Executes the previous functions
def possible_match_analysis(week_key):
  temp_results = match_distribution(week_key)
  write_test_to_workbook(temp_results, week_key)

def create_blank_txt(filename):
  with open(filename +'.txt', mode='w', encoding='utf-8') as f:
    print("New text file created!")


