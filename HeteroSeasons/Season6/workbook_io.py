#Andrew Anastasiades
#Practice Read from Workbook

import openpyxl

#Define workbook
wb = openpyxl.load_workbook(filename = "Analysis.xlsx", read_only = True, data_only = True)

#Info is the first sheet (0)
ws = wb.active


#Find the names and indices of all the guys
def get_guys():
    return_dict = dict()
    guys_range = wb.defined_names["Guys"]
    guys_dests = guys_range.destinations # returns a generator of (worksheet title, cell range) tuples

    guys_cells = []

    for title, coord in guys_dests:
        ws = wb[title]
        guys_cells.append(ws[coord])

    for each in guys_cells:
        for entry in each:
            index = entry[0]
            name = entry[1]
            return_dict[index.value] = name.value
    return return_dict

#Find the names and indices of all the girls
def get_girls():
    return_dict= dict()
    girls_range = wb.defined_names["Girls"]
    girls_dests = girls_range.destinations # returns a generator of (worksheet title, cell range) tuples

    girls_cells = []
    for title, coord in girls_dests:
        ws = wb[title]
        girls_cells.append(ws[coord])

    for each in girls_cells:
        for entry in each:
            index = entry[0]
            name = entry[1]
            return_dict[index.value] = name.value
    return return_dict

# Get all the Truth Booths
def get_booths():
    return_dict = dict()
    booth_range = wb.defined_names["Truth_Booth"]
    booth_dests = booth_range.destinations # returns a generator of (worksheet title, cell range) tuples

    booth_cells = []

    for title, coord in booth_dests:
        ws = wb[title]
        booth_cells.append(ws[coord])

    for each in booth_cells:
        for entry in each:
            week = entry[0].value
            girl = entry[1].value
            guy = entry[2].value
            index = entry[3].value
            value = entry[4].value
            is_match = entry[5].value
            return_dict[week] = {
                "week" : int(week),
                "girl" : str(girl),
                "guy" : str(guy),
                "booth" : (int(index), int(value)),
                "booth_is_match" : bool(is_match)
                }
    return return_dict

#Find the tuple of each
def get_ceremony(num_pairs, cer_name):
    return_list = list()
    cer_list = list()
    cer_range = wb.defined_names[cer_name]
    cer_dests = cer_range.destinations # returns a generator of (worksheet title, cell range) tuples

    cer_cells = []

    for title, coord in cer_dests:
        ws = wb[title]
        cer_cells.append(ws[coord])

    for each in cer_cells:
        for i in range(num_pairs):
            cer_list.append(each[i][0].value)
    matches = cer_cells[0][num_pairs][0].value
    cer_tuple = tuple(cer_list)
    return_list = [cer_tuple , matches]
    return return_list

    
